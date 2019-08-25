from bs4 import BeautifulSoup
import os
import openpyxl #https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import datetime
import time
import requests
from lxml import etree
import pdb

def update_listings(workbook):
	worksheet = workbook.active
	redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
	for i in range(2, max_row+1): # check all excel columns for this html listing
		try: 
			listingurl = worksheet.cell(row = i, column = 4).value
			page = requests.get(listingurl)
			soup = BeautifulSoup(page.text, "html.parser")
		except:
			print('Error opening listing. Please download html data and try again.')

		try:
			price = list(soup.body.find_all('script'))[15].string.split('formatted_price:')[1].strip('{text:"').split('"},max_var')[0]
		except:
			print('Error with listing',worksheet.cell(row = i, column = 1).value)
			worksheet.cell(row = i, column = 7).value = 'Error parsing prices'
			continue
		if worksheet.cell(row = i, column = 2).value != price: #price has changed so update price, add price history to last column, and color price red to notify user
			if worksheet.cell(row = i, column = 1).value != None:
				oldprice = worksheet.cell(row = i, column = 2).value
				worksheet.cell(row = i, column = 2).value = price
				if worksheet.cell(row = i, column = 6).value != None: # add to existing price changes
					worksheet.cell(row = i, column = 6).value = str(oldprice)+':'+datetime.datetime.now().strftime("%m-%d %H:%M") + ',' + str(worksheet.cell(row = i, column = 6).value)
				else: #no existing price changes
					worksheet.cell(row = i, column = 6).value = str(oldprice)+':'+datetime.datetime.now().strftime("%m-%d %H:%M")
				worksheet.cell(row = i, column = 2).fill = redFill
				print('new price found for',worksheet.cell(row = i, column = 1).value,'from',oldprice,'to',price)
		else:
			print('Price has not changed for', worksheet.cell(row = i, column = 1).value)

	workbook.save(wbname)
	os.startfile(wbname)



print('\n---------------------------------------Marketplace-Scraper--------------------------------------\n')
time.sleep(1)
print('A program to parse and save Facebook marketplace listings as well as track price changes over time.')
print('Please rerun as frequently as you would like price updates.\n')
time.sleep(1)
print('Written by David King this early morning of August the 23rd, 2019.\n')
time.sleep(1)

geolocator = Nominatim(user_agent="specify_your_app_name_here")
#home = geolocator.geocode('Dayton, OH')

#define the path to the files (with two methods)
path = 'G://My Drive/Documents/GDrivePython/'
directory = str(os.path.dirname(__file__))+'/'
wbname = 'marketbook.xlsx'

try:
	workbook = openpyxl.load_workbook(wbname)
	worksheet = workbook.active
	print('A local Excel file of marketplace data has been found!')
except:
	workbook = Workbook()
	print('No existing Excel file found. Creating a new one.')
	worksheet = workbook.active
	worksheet.append(['Title','Price','Location','Marketplace URL','Maps URL','Keywords'])
max_row = worksheet.max_row

#input_string = input("Do you want to automatically fetch updated prices in an existing Excel file? y/n:")
#if input_string == 'y':
#	update_listings(workbook)
#	exit()

print('Please download the html data from the marketplace webpage of your choice:')
print('    Scroll to the end of relevant results')
print('    Right click on a blank part of the page')
print('    Click \"save-as\"')
print('    Save the file in the same directory as this program\n')
time.sleep(1)

input_string = input("Do you want to use custom keywords? y/n:")
if input_string == 'y':
	customkeywords = True
else:
	customkeywords = False
	newkeywords = False

if customkeywords:
	try: #get keywords from the file if not given by the user
		keywordstring = worksheet.cell(row = 1, column = 7).value
		keywordlist = keywordstring.split(',')
		print('\nYour keywords saved in the local excel file are',keywordlist)
		time.sleep(1)
		print('If you do not want to change these keywords, simply press enter.')
	except:
		time.sleep(1)

	input_string = input("Enter all new keywords of interest separated by a comma: ")
	if input_string == '':
		newkeywords = False
	else:
		nonewkeywords = True
		keywordlist = input_string.split(',')
		worksheet.cell(row = 1, column = 7).value = ",".join(keywordlist)


# if customkeywords and newkeywords:
# 	worksheet.cell(row = 1, column = 7).value = ",".join(keywordlist)
# elif customkeywords:
# 	try:
# 		keywordstring = worksheet.cell(row = 1, column = 7).value
# 		keywordlist = keywordstring.split(',')
# 	except:
# 		print('There is no local excel file to store your keywords. Please restart and input keywords.')

#parce the html file
try: 
	soup = BeautifulSoup(open("marketplace.html"), "html.parser")
except:
	print('No html data located in the local directory. Please download html data and try again.')
	time.sleep(3)
	exit()
listings = soup.find_all('div',class_='_7yc _3ogd')
redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
# range through all html listings and if that html listing does not exist in the excel file, append it to the bottom
for listing in listings:
	listingfound = False
	price = listing.a.div.div.text
	title = listing.a.div.p.text
	location = listing.a.div.span.text
	# try: 
	#     geolocation = geolocator.geocode(location)
	#     distance = int(geodesic((home.latitude,home.longitude),(geolocation.latitude,geolocation.longitude)).miles)
	# except: 
	#     distance = ''
	url = listing.a['href']
	mapurl = "https://www.google.com/maps/place/"+location
	dataarray = [title, price,location,url,mapurl]
	if customkeywords: res = [ele for ele in keywordlist if (ele in title)]
	else: res = True
	#if 'ktm' in title or 'KTM' in title or 'Ktm' in title or 'Beta' in title or 'beta' in title or 'Husqvarna' in title or 'husqvarna' in title: # then we are interested in the html listing
	if bool(res):
		for i in range(2, max_row+1): # check all excel rows for this html listing
			if worksheet.cell(row = i, column = 4).value == url: # we already have the listing but lets check if the price changed
				listingfound = True
				if worksheet.cell(row = i, column = 2).value != price: #price has changes so update price, add price history to last column, and color price red to notify user
					if worksheet.cell(row = i, column = 1).value != None:
						oldprice = worksheet.cell(row = i, column = 2).value
						worksheet.cell(row = i, column = 2).value = price
						if worksheet.cell(row = i, column = 6).value != None: # add to existing price changes
							worksheet.cell(row = i, column = 6).value = str(oldprice)+':'+datetime.datetime.now().strftime("%m-%d %H:%M") + ',' + str(worksheet.cell(row = i, column = 6).value)
						else: #no existing price changes
							worksheet.cell(row = i, column = 6).value = str(oldprice)+':'+datetime.datetime.now().strftime("%m-%d %H:%M")
						worksheet.cell(row = i, column = 2).fill = redFill
						print('new price found for',title,'from',oldprice,'to',price)
		if not listingfound: # add the new listing to the excel file
			worksheet.append(dataarray)
			print('Found new listing:',title)

#Now range through cells and if there isn't a html listing that matches an excel entry, delete the excel entry
rowstodelete = []
for i in range(2, max_row+1):
	listingfound = False
	for listing in soup.find_all('div',class_='_7yc _3ogd'):
		url = listing.a['href']
		title = listing.a.div.p.text
		try:
			if customkeywords: res = [ele for ele in keywordlist if (ele in worksheet.cell(row = i, column = 1).value)]
			else: res = True
		except: pass
		if worksheet.cell(row = i, column = 4).value == url and bool(res): # an html listing matches our excel entry
			listingfound = True

	if not listingfound: # delete the excel entry with no matching html listing
		rowstodelete+=[i]
rowstodelete = rowstodelete[::-1]

for row in rowstodelete:
	deletedtitle = worksheet.cell(row = row, column = 1).value
	worksheet.delete_rows(row,1)
	print('Deleting listing:',deletedtitle)

workbook.save(wbname)
os.startfile(wbname)
